#!/usr/bin/env python3
"""
CitiVelocity CP Pattern Extractor
---------------------------------
Given a CSI ID and an authenticated Cookie header, this script:
  1. Queries /portal-admin-service/cp/query/getByCritAndPortal to enumerate
     all CPs (cpId + name) linked to that CSI ID.
  2. For each cpId, queries /portal-admin-service/cpPattern/query/queryByCpId
     to pull every URI pattern with its minAuthenticationLevel and
     exposeExternally flag.
  3. Writes the results into a single Excel workbook.

Internal use.
"""

import sys
import time
import getpass
import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL = "https://uat.citivelocity.com"
PAGE_SIZE = 50
REQUEST_TIMEOUT = 30
INTER_REQUEST_DELAY = 0.15  # seconds, gentle pacing

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:152.0) "
    "Gecko/20100101 Firefox/152.0"
)


def ts_ms() -> int:
    return int(time.time() * 1000)


def build_session(cookie: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.5",
        "Cookie": cookie,
        "Referer": f"{BASE_URL}/",
        "Connection": "keep-alive",
    })
    s.verify = False  # internal UAT often uses self-signed / MITM proxy chain
    return s


def get_cps_for_csi(session: requests.Session, csi_id: str) -> list:
    """Paginate the getByCritAndPortal endpoint until fewer than PAGE_SIZE items come back."""
    url = f"{BASE_URL}/portal-admin-service/cp/query/getByCritAndPortal"
    page = 0
    all_cps = []

    while True:
        params = {
            "pageNo": page,
            "pageSize": PAGE_SIZE,
            "portal": "CitiVelocity",
            "criteria": "csiId",
            "critValue": csi_id,
            "timeStamp": ts_ms(),
        }
        resp = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            raise RuntimeError(
                f"CP query failed (HTTP {resp.status_code}). "
                f"Body preview: {resp.text[:300]}"
            )
        try:
            data = resp.json()
        except ValueError:
            raise RuntimeError(
                "CP query returned non-JSON body (probably a login redirect). "
                "Refresh your cookie and try again."
            )

        if not isinstance(data, list):
            # Some Citi endpoints wrap the payload; try common keys
            data = data.get("data") or data.get("result") or []

        if not data:
            break

        all_cps.extend(data)
        if len(data) < PAGE_SIZE:
            break
        page += 1
        time.sleep(INTER_REQUEST_DELAY)

    return all_cps


def get_patterns_for_cp(session: requests.Session, cp_id) -> list:
    url = f"{BASE_URL}/portal-admin-service/cpPattern/query/queryByCpId"
    params = {"cpId": cp_id, "timeStamp": ts_ms()}
    resp = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
    if resp.status_code != 200:
        raise RuntimeError(
            f"Pattern query for cpId={cp_id} failed (HTTP {resp.status_code})."
        )
    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError(
            f"Pattern query for cpId={cp_id} returned non-JSON body."
        )
    if not isinstance(data, list):
        data = data.get("data") or data.get("result") or []
    return data


# ---------- Excel helpers ----------

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
RISK_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")   # exposed + weak auth
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # weak auth OR external

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True, size=11)


def write_workbook(csi_id: str, cp_records: list, out_path: str) -> None:
    """
    cp_records: list of dicts:
        {
            "cpId": int,
            "cpName": str,
            "patterns": [ {uriPattern, minAuthenticationLevel, exposeExternally, ...}, ... ],
            "error": str | None
        }
    """
    wb = Workbook()

    # ---------- Sheet 1: Details ----------
    ws = wb.active
    ws.title = "Patterns"

    ws["A1"] = "CSI ID:"
    ws["B1"] = csi_id
    ws["A1"].font = TITLE_FONT
    ws["B1"].font = TITLE_FONT

    total_cps = len(cp_records)
    total_patterns = sum(len(r["patterns"]) for r in cp_records)
    ws["A2"] = "Total CPs:"
    ws["B2"] = total_cps
    ws["A3"] = "Total Patterns:"
    ws["B3"] = total_patterns
    for c in ("A2", "A3"):
        ws[c].font = LABEL_FONT

    header_row = 5
    headers = [
        "CP ID",
        "CP Name",
        "URI Pattern",
        "Min Auth Level",
        "Expose Externally",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    row = header_row + 1
    for rec in cp_records:
        cp_id = rec["cpId"]
        cp_name = rec["cpName"]

        if rec.get("error"):
            ws.cell(row=row, column=1, value=cp_id).border = BORDER
            ws.cell(row=row, column=2, value=cp_name).border = BORDER
            ws.cell(row=row, column=3, value=f"ERROR: {rec['error']}").border = BORDER
            row += 1
            continue

        if not rec["patterns"]:
            ws.cell(row=row, column=1, value=cp_id).border = BORDER
            ws.cell(row=row, column=2, value=cp_name).border = BORDER
            ws.cell(row=row, column=3, value="(no patterns)").border = BORDER
            row += 1
            continue

        for p in rec["patterns"]:
            uri = p.get("uriPattern")
            min_auth = p.get("minAuthenticationLevel")
            expose = p.get("exposeExternally")

            row_cells = [
                ws.cell(row=row, column=1, value=cp_id),
                ws.cell(row=row, column=2, value=cp_name),
                ws.cell(row=row, column=3, value=uri),
                ws.cell(row=row, column=4, value=min_auth),
                ws.cell(row=row, column=5, value=expose),
            ]

            # Highlight security-relevant combinations
            weak_auth = str(min_auth).upper() in ("NONE", "0", "N", "NULL", "")
            externally_exposed = str(expose).upper() in ("Y", "YES", "TRUE")

            if weak_auth and externally_exposed:
                fill = RISK_FILL
            elif weak_auth or externally_exposed:
                fill = WARN_FILL
            else:
                fill = None

            for c in row_cells:
                c.border = BORDER
                if fill:
                    c.fill = fill

            row += 1

    widths = [12, 32, 70, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    # ---------- Sheet 2: CP Summary ----------
    ws2 = wb.create_sheet(title="CP Summary")
    ws2["A1"] = "CSI ID:"
    ws2["B1"] = csi_id
    ws2["A1"].font = TITLE_FONT
    ws2["B1"].font = TITLE_FONT

    sum_headers = ["CP ID", "CP Name", "Pattern Count"]
    for col_idx, h in enumerate(sum_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    r = 4
    for rec in cp_records:
        ws2.cell(row=r, column=1, value=rec["cpId"]).border = BORDER
        ws2.cell(row=r, column=2, value=rec["cpName"]).border = BORDER
        count_cell = ws2.cell(
            row=r,
            column=3,
            value=("ERR" if rec.get("error") else len(rec["patterns"])),
        )
        count_cell.border = BORDER
        r += 1

    for i, w in enumerate([12, 40, 15], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    # ---------- Sheet 3: Legend ----------
    ws3 = wb.create_sheet(title="Legend")
    ws3["A1"] = "Row Highlighting"
    ws3["A1"].font = TITLE_FONT
    ws3["A3"] = "Orange"
    ws3["A3"].fill = RISK_FILL
    ws3["B3"] = "Exposed externally AND minAuthenticationLevel is weak/NONE"
    ws3["A4"] = "Yellow"
    ws3["A4"].fill = WARN_FILL
    ws3["B4"] = "Exposed externally OR minAuthenticationLevel is weak/NONE"
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 80

    wb.save(out_path)


# ---------- CLI ----------

def prompt_multiline_cookie() -> str:
    """
    Cookies can be very long. Accept either a single-line paste or a
    multi-line paste terminated by an empty line.
    """
    print("Paste the full Cookie header value.")
    print("  - Single line paste + Enter is fine.")
    print("  - Or paste multiple lines, then press Enter on an empty line to finish.")
    first = input("Cookie: ").strip()
    if not first:
        print("[!] Empty cookie, aborting.")
        sys.exit(1)

    # If the user just pasted a single line, use it as-is.
    # Otherwise let them add continuation lines.
    lines = [first]
    try:
        while True:
            more = input()
            if not more.strip():
                break
            lines.append(more.strip())
    except EOFError:
        pass

    return " ".join(lines)


def main():
    print("=" * 70)
    print("  CitiVelocity CP Pattern Extractor")
    print("=" * 70)

    csi_id = input("Enter CSI ID (e.g. 171632): ").strip()
    if not csi_id:
        print("[!] CSI ID is required.")
        sys.exit(1)

    cookie = prompt_multiline_cookie()

    session = build_session(cookie)

    print(f"\n[*] Fetching CPs for CSI ID {csi_id} ...")
    try:
        cps = get_cps_for_csi(session, csi_id)
    except Exception as e:
        print(f"[!] {e}")
        sys.exit(1)

    if not cps:
        print("[!] No CPs returned. Check the CSI ID or the cookie.")
        sys.exit(1)

    print(f"[+] Found {len(cps)} CP(s):")
    for cp in cps:
        print(f"    - cpId={cp.get('cpId')}  name={cp.get('name')}")

    cp_records = []
    for i, cp in enumerate(cps, start=1):
        cp_id = cp.get("cpId")
        cp_name = cp.get("name")
        print(f"\n[*] ({i}/{len(cps)}) Fetching patterns for cpId={cp_id} ({cp_name})")

        rec = {"cpId": cp_id, "cpName": cp_name, "patterns": [], "error": None}
        try:
            patterns = get_patterns_for_cp(session, cp_id)
            rec["patterns"] = patterns
            print(f"    [+] {len(patterns)} pattern(s)")
        except Exception as e:
            rec["error"] = str(e)
            print(f"    [!] {e}")

        cp_records.append(rec)
        time.sleep(INTER_REQUEST_DELAY)

    out_path = f"citivelocity_csi_{csi_id}_{int(time.time())}.xlsx"
    write_workbook(csi_id, cp_records, out_path)
    print(f"\n[+] Done. Excel saved to: {out_path}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[!] Interrupted by user.")
        sys.exit(130)
